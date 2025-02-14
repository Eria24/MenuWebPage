from flask import Flask, request, jsonify
from flask_cors import CORS  # Import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Path to the Excel file
EXCEL_FILE = "menu_selections.xlsx"

@app.route("/save-selections", methods=["POST"])
def save_selections():
    # Get the selected items from the request
    data = request.json
    selections = data.get("selections")

    # Check if the Excel file exists
    if not os.path.exists(EXCEL_FILE):
        # Create a new Excel file with headers if it doesn't exist
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Timestamp", "Selections"])  # Add headers
    else:
        # Load the existing Excel file
        workbook = load_workbook(EXCEL_FILE)
        sheet = workbook.active

    # Add the new selection with a timestamp
    timestamp = datetime.now().isoformat()
    sheet.append([timestamp, selections])

    # Save the updated Excel file
    workbook.save(EXCEL_FILE)

    # Return a success response
    return jsonify({"success": True})

if __name__ == "__main__":
    app.run(debug=True)